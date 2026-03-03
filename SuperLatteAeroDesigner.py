# -*- coding: utf-8 -*-
"""
LatteAeroDesigner.py (fixed)

1) 設計Excel(BASE / BASE限界)の翼型定義と迎角分布を参照
2) CSVフォルダの極データから「近傍2Re」を自動選択し、Re方向に線形補間
3) LLTで誘導迎角分布を計算して「定常時揚力分布/低速時揚力分布」へ書き戻し
4) 誘導迎角で補正した有効迎角 alpha_eff を用いて、Cl/Cd/Cm/Xcp を書き込み

今回の修正点（重要）
- 「定常時揚力分布」の chord が空でも落ちない：
  chord が空なら BASE の chord 定義からその場で補完して処理を継続
- chord 定義（BASE!F/H）の欠落があっても、利用可能な節点から chord(y) を作る：
  L14..L19 の境界点で得られる chord 節点を集め、y方向に線形補間（端はクランプ）
- 翼端より先は chord/誘導迎角/係数を空欄にする（従来通り）
"""

from __future__ import annotations

import os
import re
import math
from dataclasses import dataclass
from typing import Optional, List, Any, Tuple, Dict

import numpy as np
import pandas as pd
import unicodedata as ud
from openpyxl import load_workbook


# =========================
# 設定（ここだけ触ればOK）
# =========================
WORKBOOK_PATH = r"C:\aero\末端技研機設計シート.xlsx"
CSV_FOLDER    = r"C:\aero\csv"

BASE_SHEET   = "BASE"
BASE_SHEET2  = "BASE限界"


DEBUG = False

# =========================
# Re 自動計算（I/K を参照しない）
# =========================
# ※画像の配置に合わせたデフォルト。あなたの設計シートで違う場合はここを変更。
V_CELL  = "F4"   # BASE!F4  : 定常機速 V∞ [m/s]
NU_CELL = "F10"  # BASE!F10 : 動粘性係数 ν [m^2/s]


# =========================
# 「定常時揚力分布」反映設定
# =========================
LIFT_SHEET       = "定常時揚力分布"
LIFT_SHEET2      = "低速時揚力分布"
LIFT_Y_COL       = 1   # A列: 翼根から距離(mm)
LIFT_A_START_ROW = 4   # ここから下にデータが並ぶ
INDUCED_COL      = 4   # D列: 誘導迎角(deg)を書き込む

CHORD_COL        = 2   # ★仮定: B列に chord(mm)。違うなら変更

# 出力列（指定）
OUT_CL_COL  = 5
OUT_CD_COL  = 6
OUT_CM_COL  = 7
OUT_XCP_COL = 8

# 翼効率 e の書き込みセル
E_CELL = "N7"  # BASE!N7

# a0, alpha0 推定用の回帰範囲（deg）
FIT_ALPHA_MIN = -4.0
FIT_ALPHA_MAX = +4.0


# =========================
# A列スキップ判定（XFLR5ヘッダ等）
# =========================
SKIP_EQ_STRINGS = [
    "1 1 Reynolds number fixed Mach number fixed",
    "xflr5 v6.61",
    " xtrf =   1.000 (top)        1.000 (bottom)"
]

SKIP_CONTAINS_ALL = [
    ["reynolds number fixed", "mach number fixed", "Calculated polar for:", "xtrf", "Mach"],
]

SKIP_REGEXES = [
    r"^\s*#.*$",
    r"\b(convergence|failed)\b",
]


# =========================
# ユーティリティ
# =========================
def _clean_for_filename(x: Any) -> str:
    """Excelセルの翼型名等をファイル名用に整形"""
    if x is None:
        return ""
    if isinstance(x, float) and x.is_integer():
        return str(int(x))
    return str(x).strip()

def _numify(s: Any) -> Optional[float]:
    """頑健な数値抽出（全角/NBSP/度記号/カンマ等）"""
    if s is None:
        return None
    t = str(s)
    t = t.replace("\u00a0", " ")
    t = ud.normalize("NFKC", t)
    t = t.replace("°", "").replace("deg", "").replace("DEG", "")
    t = t.strip()
    if "," in t and "." not in t:
        t = t.replace(",", ".")
    else:
        t = t.replace(",", "")
    m = re.search(r'[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?', t)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None

def _clean_str(s: Any) -> str:
    """数値化できない場合の文字列比較用"""
    if s is None:
        return ""
    t = ud.normalize("NFKC", str(s)).replace("\u00a0", " ")
    t = t.replace("°", "").replace("deg", "").replace("DEG", "")
    return t.strip()

def _norm_collapse(s: Any) -> str:
    t = ud.normalize("NFKC", str(s)).replace("\u00a0", " ")
    t = re.sub(r"\s+", " ", t).strip()
    return t

def _is_skip_string(s: Any) -> bool:
    if s is None:
        return False
    t_norm = _norm_collapse(s)
    t_low  = t_norm.lower()

    for x in SKIP_EQ_STRINGS:
        if t_norm == _norm_collapse(x):
            return True

    for group in SKIP_CONTAINS_ALL:
        if all(g.lower() in t_low for g in group):
            return True

    for pat in SKIP_REGEXES:
        if re.search(pat, t_norm, flags=re.IGNORECASE):
            return True

    return False

def _is_empty_or_zero(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, (int, float)) and float(x) == 0.0:
        return True
    s = str(x).strip()
    return s == "" or s == "0" or s == "0.0"

def write_row_to_sheet(ws, target_row: int, values: List[Any]) -> None:
    for i, v in enumerate(values, start=1):
        ws.cell(row=target_row, column=i).value = v


# =========================
# Re計算ユーティリティ
# =========================
def _get_vinf_nu(ws_base) -> Tuple[float, float]:
    V = _numify(ws_base[V_CELL].value)
    nu = _numify(ws_base[NU_CELL].value)
    if V is None or V <= 0:
        raise ValueError(f"{ws_base.title}!{V_CELL} のV∞が不正です: {ws_base[V_CELL].value}")
    if nu is None or nu <= 0:
        raise ValueError(f"{ws_base.title}!{NU_CELL} のνが不正です: {ws_base[NU_CELL].value}")
    return float(V), float(nu)

def _re_from_chord_mm(ws_base, chord_mm: float) -> float:
    """Re = V * c / nu（cはmm入力）"""
    V, nu = _get_vinf_nu(ws_base)
    return float(V * (float(chord_mm) / 1000.0) / nu)

def _segment_bounds_y(ws_base, row: int) -> Tuple[float, float]:
    """row(14..19)の区間 [y0,y1] を mm で返す"""
    y0 = 0.0
    for r in range(14, row):
        L = _numify(ws_base[f"L{r}"].value)
        if L is None:
            continue
        y0 += float(L)
    Lr = _numify(ws_base[f"L{row}"].value)
    if Lr is None or Lr <= 0:
        raise ValueError(f"{ws_base.title}!L{row} が不正です: {ws_base[f'L{row}'].value}")
    y1 = y0 + float(Lr)
    return float(y0), float(y1)

def _chord_from_lift_sheet(ws_lift, y_mm: float, y_tip_mm: float) -> float:
    """揚力分布シート(A:y, B:chord)から chord(y) を補間取得。"""
    ys: List[float] = []
    cs: List[float] = []
    r = LIFT_A_START_ROW
    while True:
        yv = ws_lift.cell(row=r, column=LIFT_Y_COL).value
        if yv is None or str(yv).strip() == "":
            break
        y = _numify(yv)
        c = _numify(ws_lift.cell(row=r, column=CHORD_COL).value)
        if (y is not None) and (c is not None) and (float(y) <= y_tip_mm + 1e-9):
            ys.append(float(y))
            cs.append(float(c))
        r += 1

    if len(ys) < 2:
        raise ValueError(f"{ws_lift.title} から chord を補間できません（点数不足）。")

    ys_np = np.array(ys, dtype=float)
    cs_np = np.array(cs, dtype=float)
    order = np.argsort(ys_np)
    ys_np = ys_np[order]
    cs_np = cs_np[order]

    yq = float(y_mm)
    if yq <= float(ys_np[0]):
        return float(cs_np[0])
    if yq >= float(ys_np[-1]):
        return float(cs_np[-1])
    return float(np.interp(yq, ys_np, cs_np))


# =========================
# CSV読み込み（エンコ/区切り推定）
# =========================
def _try_read_csv_hard(csv_path: str) -> Optional[pd.DataFrame]:
    encodings = ["utf-8-sig", "utf-8", "cp932", "shift_jis", "utf-16", "utf-16le", "utf-16be", "latin1"]
    seps = [None, ",", "\t", ";", "|", r"\s+"]
    headers = [None, 0]
    for enc in encodings:
        for sep in seps:
            for header in headers:
                try:
                    kwargs = dict(header=header, dtype=str, encoding=enc, engine="python")
                    kwargs["sep"] = None if sep is None else sep
                    df = pd.read_csv(csv_path, **kwargs)
                    if df is not None and not df.empty:
                        return df
                except Exception:
                    continue
    return None

def _expand_single_column_df(df: pd.DataFrame) -> pd.DataFrame:
    """CSVが1列として読まれてしまった場合、カンマ or 空白で強制分割する"""
    if df is None or df.empty:
        return df
    if df.shape[1] != 1:
        return df


def _read_xflr5_polar_table(csv_path: str) -> pd.DataFrame:
    # XFLR5のpolar CSVを確実に読む（先頭の説明行をスキップし、数値行のみ解析）
    from pathlib import Path as _Path
    text = _Path(csv_path).read_text(encoding="utf-8", errors="replace").splitlines()

    header = None
    header_idx = None
    for i, line in enumerate(text):
        t = line.strip()
        if not t:
            continue
        tl = t.lower().replace(" ", "")
        if tl.startswith("alpha,") and ("cl" in tl) and ("cd" in tl):
            header = [h.strip() for h in t.split(",")]
            header_idx = i
            break

    # ヘッダが見つからない場合は従来ローダへフォールバック
    if header is None or header_idx is None:
        df = _try_read_csv_hard(csv_path)
        if df is None:
            raise ValueError(f"CSV読込に失敗: {csv_path}")
        return _expand_single_column_df(df)

    rows = []
    maxlen = len(header)
    for line in text[header_idx + 1:]:
        t = line.strip()
        if (not t) or ("," not in t):
            continue
        if not re.match(r"^[-+]?(?:\d|\.\d)", t):  # alphaが先頭の数値行のみ
            continue
        parts = [p.strip() for p in t.split(",")]
        if len(parts) > maxlen:
            maxlen = len(parts)
        rows.append(parts)

    if not rows:
        raise ValueError(f"データ行が見つかりません: {csv_path}")

    cols = header[:]
    if len(cols) < maxlen:
        cols += [f"col{j}" for j in range(len(cols), maxlen)]

    rows = [r + [""] * (maxlen - len(r)) for r in rows]
    return pd.DataFrame(rows, columns=cols)

    s = df.iloc[:, 0].astype(str)

    split_comma = s.str.split(",", expand=True)
    if split_comma.shape[1] >= 2:
        return split_comma

    split_ws = s.str.strip().str.split(r"\s+", expand=True)
    if split_ws.shape[1] >= 2:
        return split_ws

    return df


# =========================
# A列一致検索（AoAなど）
# =========================
def _find_match_index_in_Acol(df: pd.DataFrame, needle: Any) -> Optional[int]:
    a_col = df.iloc[:, 0]
    skip_mask = a_col.map(_is_skip_string)
    if skip_mask.any():
        a_col = a_col[~skip_mask]

    col_num = pd.Series([_numify(v) for v in a_col], index=a_col.index)
    n_num = _numify(needle)
    if n_num is not None:
        col_arr = np.array([np.nan if v is None else float(v) for v in col_num.to_numpy()], dtype=float)
        mask = np.isfinite(col_arr) & np.isclose(col_arr, float(n_num), rtol=1e-9, atol=1e-7)
        if mask.any():
            first_pos = int(np.argmax(mask))
            return col_num.index[first_pos]

    a_clean = a_col.map(_clean_str)
    n_clean = _clean_str(needle)
    eq_mask = (a_clean == n_clean).to_numpy()
    if eq_mask.any():
        first_pos = int(np.argmax(eq_mask))
        return a_clean.index[first_pos]

    return None


# =========================
# Re近傍2点探索（ファイル名からRe抽出）
# =========================
def _extract_re_from_filename(filename: str) -> Optional[float]:
    m = re.search(r"_T1_Re0\.([0-9]+(?:\.[0-9]+)?)_", filename)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None

def _find_two_nearest_re_files(token: str, re_target: float) -> Tuple[str, str, float, float, float]:
    if not os.path.isdir(CSV_FOLDER):
        raise NotADirectoryError(f"CSV フォルダが存在しません: {CSV_FOLDER}")

    candidates = [f for f in os.listdir(CSV_FOLDER)
                  if f.lower().endswith(".csv") and f.startswith(f"{token}_T1_Re0.")]
    if not candidates:
        raise FileNotFoundError(f"token={token} のCSVが見つかりません。")

    pairs = []
    for f in candidates:
        rv = _extract_re_from_filename(f)
        if rv is None:
            continue
        pairs.append((rv, f))
    if not pairs:
        raise FileNotFoundError(f"token={token} のCSVはあるが、Re抽出に失敗しました（命名規則が想定外）。")

    pairs.sort(key=lambda x: x[0])
    re_list = [p[0] for p in pairs]

    if len(pairs) == 1:
        re1, f1 = pairs[0]
        p = os.path.join(CSV_FOLDER, f1)
        return p, p, re1, re1, 0.0

    idx = int(np.searchsorted(re_list, re_target))
    if idx <= 0:
        re1, f1 = pairs[0]
        re2, f2 = pairs[1]
    elif idx >= len(pairs):
        re1, f1 = pairs[-2]
        re2, f2 = pairs[-1]
    else:
        re1, f1 = pairs[idx - 1]
        re2, f2 = pairs[idx]

    if re2 == re1:
        alpha = 0.0
    else:
        alpha = (re_target - re1) / (re2 - re1)
        alpha = float(max(0.0, min(1.0, alpha)))  # 外挿しない

    return (os.path.join(CSV_FOLDER, f1),
            os.path.join(CSV_FOLDER, f2),
            float(re1), float(re2), float(alpha))


# =========================
# 行のRe補間（セル値）
# =========================
def _interp_row(v1: Any, v2: Any, alpha: float) -> Any:
    n1 = _numify(v1)
    n2 = _numify(v2)
    if n1 is not None and n2 is not None:
        return float(n1 + alpha * (n2 - n1))
    return v1 if alpha < 0.5 else v2


# =========================
# LLT（区間割り当て込み） + Excel書き戻し
# =========================
@dataclass
class Segment:
    y0: float      # mm
    y1: float      # mm
    foil0: str
    foil1: str
    re0: float
    re1: float
    a0: float      # deg
    a1: float      # deg

def _lerp(x0: float, x1: float, t: float) -> float:
    return float(x0 + t * (x1 - x0))


def _build_segments_from_BASE(ws_base) -> List[Segment]:
    """
    旧：BASEの I/K (Re) を参照
    新：V∞, ν と chord端(F/H)から Re を内部計算
         chord端が欠落している場合は chord節点から補間
    """
    segs: List[Segment] = []
    y = 0.0

    # chord節点（欠落補完用）
    yk, ck, _ = _build_chord_knots_from_BASE(ws_base)

    for r in range(14, 19):  # 14..18（LLT区間）
        L = _numify(ws_base[f"L{r}"].value)
        if L is None or L <= 0:
            continue

        foil0 = _clean_for_filename(ws_base[f"C{r}"].value)
        foil1 = _clean_for_filename(ws_base[f"E{r}"].value)

        y0 = float(y)
        y1 = float(y + L)

        c0 = _numify(ws_base[f"F{r}"].value)
        c1 = _numify(ws_base[f"H{r}"].value)
        if c0 is None:
            c0 = _chord_at_from_knots(yk, ck, y0)
        if c1 is None:
            c1 = _chord_at_from_knots(yk, ck, y1)

        re0 = _re_from_chord_mm(ws_base, float(c0))
        re1 = _re_from_chord_mm(ws_base, float(c1))

        a0 = float(ws_base[f"N{r}"].value)
        a1 = float(ws_base[f"P{r}"].value)

        segs.append(Segment(y0=y0, y1=y1, foil0=foil0, foil1=foil1, re0=re0, re1=re1, a0=a0, a1=a1))
        y = y1

    return segs

def _pick_segment(segs: List[Segment], y_mm: float) -> Tuple[Segment, float]:
    if y_mm <= segs[0].y0:
        return segs[0], 0.0
    for s in segs:
        if y_mm <= s.y1 + 1e-9:
            t = 0.0 if s.y1 == s.y0 else (y_mm - s.y0) / (s.y1 - s.y0)
            return s, float(max(0.0, min(1.0, t)))
    return segs[-1], 1.0


def _read_polar_df(token: str, re_target: float) -> pd.DataFrame:
    p1, p2, _, _, a = _find_two_nearest_re_files(token, re_target)

    df1 = _read_xflr5_polar_table(p1)
    df2 = _read_xflr5_polar_table(p2)

    alpha1 = pd.to_numeric(df1.iloc[:, 0].map(_numify), errors="coerce")
    alpha2 = pd.to_numeric(df2.iloc[:, 0].map(_numify), errors="coerce")

    df1n = df1.copy()
    df2n = df2.copy()
    df1n.iloc[:, 0] = alpha1
    df2n.iloc[:, 0] = alpha2

    key = df1n.columns[0]
    df1n = df1n.dropna(subset=[key]).sort_values(key)
    df2n = df2n.dropna(subset=[key]).sort_values(key)

    m = pd.merge(df1n, df2n, on=key, suffixes=("_1", "_2"), how="inner")
    if m.empty:
        raise ValueError(f"{token} の Re1/Re2で共通alphaが無く、Re補間できません。")

    out = pd.DataFrame()
    out[key] = m[key]

    for j in range(1, min(df1n.shape[1], df2n.shape[1])):
        c1 = df1n.columns[j]
        s1 = pd.to_numeric(m[f"{c1}_1"].map(_numify), errors="coerce")
        s2 = pd.to_numeric(m[f"{c1}_2"].map(_numify), errors="coerce")
        out[c1] = s1 + a * (s2 - s1)

    return out.reset_index(drop=True)

def _fit_a0_alpha0(polar: pd.DataFrame) -> Tuple[float, float]:
    alpha_deg = pd.to_numeric(polar.iloc[:, 0], errors="coerce")
    cl = pd.to_numeric(polar.iloc[:, 1], errors="coerce")

    mask = alpha_deg.between(FIT_ALPHA_MIN, FIT_ALPHA_MAX) & np.isfinite(cl)
    a = alpha_deg[mask].to_numpy()
    y = cl[mask].to_numpy()
    if len(a) < 3:
        raise ValueError("a0/alpha0 推定点が不足（FIT_ALPHA_MIN/MAXを調整）。")

    a_rad = np.deg2rad(a)
    m, b = np.polyfit(a_rad, y, 1)
    alpha0 = -b / m
    a0 = m
    return float(a0), float(alpha0)

def _lookup_coeffs_from_polar(polar: pd.DataFrame, alpha_deg_target: float) -> Tuple[float, float, float, float]:
    vals = polar.copy()
    vals.iloc[:, 0] = pd.to_numeric(vals.iloc[:, 0], errors="coerce")
    vals = vals.dropna(subset=[vals.columns[0]]).sort_values(vals.columns[0])

    a = vals.iloc[:, 0].to_numpy()
    if len(a) < 2:
        raise ValueError("polar の点数が少なすぎます。")

    cl = pd.to_numeric(vals.iloc[:, 1], errors="coerce").to_numpy()
    cd = pd.to_numeric(vals.iloc[:, 2], errors="coerce").to_numpy() if vals.shape[1] > 2 else np.full_like(cl, np.nan)
    cm = pd.to_numeric(vals.iloc[:, 4], errors="coerce").to_numpy() if vals.shape[1] > 4 else np.full_like(cl, np.nan)
    xcp = pd.to_numeric(vals.iloc[:, 11], errors="coerce").to_numpy() if vals.shape[1] > 11 else np.full_like(cl, np.nan)

    at = float(alpha_deg_target)
    if at <= a[0]:
        i0, i1 = 0, 1
    elif at >= a[-1]:
        i0, i1 = -2, -1
    else:
        i1 = int(np.searchsorted(a, at))
        i0 = i1 - 1

    a0, a1 = a[i0], a[i1]
    t = 0.0 if a1 == a0 else (at - a0) / (a1 - a0)
    # 範囲外での暴走を防ぐ（外挿しない）
    t = float(max(0.0, min(1.0, t)))

    def lerp(u, v):
        return float(u + t * (v - u))

    return lerp(cl[i0], cl[i1]), lerp(cd[i0], cd[i1]), lerp(cm[i0], cm[i1]), lerp(xcp[i0], xcp[i1])


_POLAR_CACHE: Dict[Tuple[str, float], pd.DataFrame] = {}
_FIT_CACHE: Dict[Tuple[str, float], Tuple[float, float]] = {}

def _get_polar_cached(token: str, re_target: float) -> pd.DataFrame:
    key = (token, round(float(re_target), 3))
    if key not in _POLAR_CACHE:
        _POLAR_CACHE[key] = _read_polar_df(token, float(re_target))
    return _POLAR_CACHE[key]

def _get_a0_alpha0_cached(token: str, re_target: float) -> Tuple[float, float]:
    key = (token, round(float(re_target), 3))
    if key not in _FIT_CACHE:
        polar = _get_polar_cached(token, re_target)
        _FIT_CACHE[key] = _fit_a0_alpha0(polar)
    return _FIT_CACHE[key]


def _llt_solve_halfspan(y_m: np.ndarray, c_m: np.ndarray,
                        a0_1perrad: np.ndarray,
                        alpha_geo_rad: np.ndarray,
                        alpha0_rad: np.ndarray) -> Tuple[np.ndarray, np.ndarray, float]:
    """
    半翼・対称翼 専用の安定化LLT

    - 解析は半翼のみ（y: 0..b/2）
    - 循環のフーリエ展開は「奇数次のみ」（対称翼 → 偶数次は0）を用いる
        n = 1,3,5,..., (2N-1)
    - 連立方程式（標準LLT）：
        Σ A_n [ (4b/(a0 c)) sin(nθ) + n sin(nθ)/sinθ ] = (α_geo - α0)
      ここで b は全スパン（=2*b/2）

    戻り値:
      y_out_m : root→tip の y[m]
      alpha_i_rad_out : root→tip の誘導迎角[rad]
      e : 翼効率（A1^2 / Σ n A_n^2）
    """
    # ---- 入力は半翼 0..b/2 を想定 ----
    b2 = float(np.max(y_m))           # b/2 [m]
    if b2 <= 0:
        raise ValueError("LLT: b/2 が 0 以下です（y_m を確認）。")
    b = 2.0 * b2                      # 全スパン [m]

    Np = int(len(y_m))                # 入力点数（後で補間に使う）
    if Np < 4:
        raise ValueError("LLT点数が少なすぎます（stationsを増やしてください）。")

    # ---- コロケーション点（θ: (0, π/2) 内部点）----
    # tip(θ→0) と root(θ→π/2) を避ける
    Nc = Np  # コロケーション点数（入力点数と同じでOK）
    # tip/root から離す mid-point 配置
    theta = np.array([(i + 0.5) * (0.5 * np.pi) / Nc for i in range(Nc)], dtype=float)  # mid-point
    y_theta = b2 * np.cos(theta)      # tip→root の並び

    # ---- yでソートして θ点へ補間 ----
    order = np.argsort(y_m)
    y_sorted = y_m[order]
    c_sorted = c_m[order]
    a0_sorted = a0_1perrad[order]
    ageo_sorted = alpha_geo_rad[order]
    a0z_sorted = alpha0_rad[order]

    def interp(arr):
        return np.interp(y_theta, y_sorted, arr)

    c = interp(c_sorted)
    a0 = interp(a0_sorted)
    ageo = interp(ageo_sorted)
    a0z = interp(a0z_sorted)
    rhs = (ageo - a0z)

    # --- Sanity checks (debug) ---
    if (not np.all(np.isfinite(c))) or (not np.all(np.isfinite(a0))) or (not np.all(np.isfinite(rhs))):
        raise ValueError("LLT入力にNaN/infがあります: chord/a0/rhs を確認してください。")
    if float(np.min(c)) <= 0.0:
        raise ValueError(f"LLT chord<=0 を検出: c_min={float(np.min(c))}")
    if float(np.min(a0)) <= 0.0:
        raise ValueError(f"LLT a0<=0 を検出: a0_min={float(np.min(a0))}")

    # ---- 奇数次のみ（n=1,3,5,...）で解く ----
    n_list = np.array([2*k - 1 for k in range(1, Nc + 1)], dtype=float)  # shape (Nc,)
    M = np.zeros((Nc, Nc), dtype=float)

    for i in range(Nc):
        si = math.sin(theta[i])
        # si は 0 に近いほど条件悪化するので、下限を設ける（内部点だが念のため）
        if si < 1e-6:
            si = 1e-6
        for j in range(Nc):
            n = n_list[j]
            sn = math.sin(n * theta[i])
            M[i, j] = (4.0 * b / (a0[i] * c[i])) * sn + (n * sn) / si

    # 連立を解く（条件が悪い場合もあるので例外を分かりやすく）
    try:
        A = np.linalg.solve(M, rhs)
    except np.linalg.LinAlgError as e:
        cond = float(np.linalg.cond(M))
        raise ValueError(f"LLT連立が解けません（特異/悪条件）。cond(M)={cond:.3e}") from e

    # ---- 誘導迎角 α_i(θ) = Σ n A_n sin(nθ)/sinθ ----
    alpha_i = np.zeros(Nc, dtype=float)
    for i in range(Nc):
        si = math.sin(theta[i])
        if si < 1e-6:
            si = 1e-6
        alpha_i[i] = float(np.sum(n_list * A * np.sin(n_list * theta[i])) / si)

    # ---- 翼効率 e = A1^2 / Σ (n A_n^2) ----
    A1 = float(A[0])
    denom = float(np.sum(n_list * (A ** 2)))
    e = float((A1 ** 2) / denom) if denom > 0 else float("nan")

    # root→tip へ並び替え（y_theta は tip→root）
    y_out = y_theta[::-1]
    alpha_i_out = alpha_i[::-1]
    return y_out, alpha_i_out, e


# =========================
# chord(y) 構築（欠落に強い）
# =========================
def _build_chord_knots_from_BASE(ws_base) -> Tuple[np.ndarray, np.ndarray, float]:
    knots: Dict[float, List[float]] = {}
    y = 0.0
    for r in range(14, 20):  # 14..19
        L = _numify(ws_base[f"L{r}"].value)
        if L is None or L <= 0:
            continue
        y0 = float(y)
        y1 = float(y + L)

        c0 = _numify(ws_base[f"F{r}"].value)
        c1 = _numify(ws_base[f"H{r}"].value)

        if c0 is not None:
            knots.setdefault(y0, []).append(float(c0))
        if c1 is not None:
            knots.setdefault(y1, []).append(float(c1))

        y = y1

    if not knots:
        raise ValueError("BASE!L14:L19 と BASE!F/H から chord 節点が1つも取れません。")

    yk = np.array(sorted(knots.keys()), dtype=float)
    ck = np.array([float(np.mean(knots[yy])) for yy in yk], dtype=float)
    y_tip = float(np.max(yk))
    return yk, ck, y_tip

def _chord_at_from_knots(yk_mm: np.ndarray, ck_mm: np.ndarray, y_mm: float) -> float:
    y = float(y_mm)
    if y <= float(yk_mm[0]):
        return float(ck_mm[0])
    if y >= float(yk_mm[-1]):
        return float(ck_mm[-1])
    return float(np.interp(y, yk_mm, ck_mm))


def write_chord_to_lift_sheet(wb, base_sheet_name: str, lift_sheet_name: str) -> None:
    ws_base = wb[base_sheet_name]
    ws_lift = wb[lift_sheet_name]

    yk, ck, y_tip_mm = _build_chord_knots_from_BASE(ws_base)

    r = LIFT_A_START_ROW
    written = 0
    while True:
        yv = ws_lift.cell(row=r, column=LIFT_Y_COL).value
        if yv is None or str(yv).strip() == "":
            break
        y_mm = float(_numify(yv) if _numify(yv) is not None else yv)

        if y_mm <= y_tip_mm + 1e-9:
            ws_lift.cell(row=r, column=CHORD_COL).value = _chord_at_from_knots(yk, ck, y_mm)
            written += 1
        else:
            ws_lift.cell(row=r, column=CHORD_COL).value = None
        r += 1

    col_letter = chr(ord('A') + CHORD_COL - 1)
    print(f"[CHORD] ({base_sheet_name} -> {lift_sheet_name}) wrote chord(mm) -> "
          f"{lift_sheet_name}!{col_letter}{LIFT_A_START_ROW}..  (n={written}, y_tip={y_tip_mm:.1f}mm)")


def llt_and_write_back(wb, base_sheet_name: str, lift_sheet_name: str, e_cell: str = "N7") -> None:
    ws_base = wb[base_sheet_name]
    ws_lift = wb[lift_sheet_name]

    segs = _build_segments_from_BASE(ws_base)
    if not segs:
        raise ValueError(f"{base_sheet_name}!L14..L18 からLLT区間が作れません。")

    yk, ck, y_tip_mm_chord = _build_chord_knots_from_BASE(ws_base)
    y_tip_mm = float(segs[-1].y1)

    y_all: List[float] = []
    c_all: List[Optional[float]] = []

    r = LIFT_A_START_ROW
    while True:
        yv = ws_lift.cell(row=r, column=LIFT_Y_COL).value
        if yv is None or str(yv).strip() == "":
            break
        y_mm = float(_numify(yv) if _numify(yv) is not None else yv)
        y_all.append(y_mm)

        cv = ws_lift.cell(row=r, column=CHORD_COL).value
        c_mm = _numify(cv)

        if (c_mm is None) and (y_mm <= y_tip_mm + 1e-9):
            c_mm = _chord_at_from_knots(yk, ck, y_mm)
            ws_lift.cell(row=r, column=CHORD_COL).value = float(c_mm)

        c_all.append(None if c_mm is None else float(c_mm))
        r += 1

    if len(y_all) < 4:
        raise ValueError(f"{lift_sheet_name} A列の点数が少なすぎます。")

    y_all_mm = np.array(y_all, dtype=float)
    c_all_mm = np.array([np.nan if v is None else float(v) for v in c_all], dtype=float)

    mask = (y_all_mm <= y_tip_mm + 1e-9) & np.isfinite(c_all_mm)
    n_use = int(mask.sum())
    if n_use < 4:
        raise ValueError(
            f"LLT対象点が少なすぎます。 y_tip(LLT)={y_tip_mm:.1f}mm, points(use)={n_use}, "
            f"y_tip(chord_knots)={y_tip_mm_chord:.1f}mm"
        )

    idx_use = np.flatnonzero(mask)
    idx_map = {int(k): i for i, k in enumerate(idx_use)}

    y_mm = y_all_mm[mask]
    c_mm = c_all_mm[mask]

    y_m = y_mm / 1000.0
    c_m = c_mm / 1000.0

    a0_arr = np.zeros_like(y_m, dtype=float)
    alpha0_arr = np.zeros_like(y_m, dtype=float)
    alpha_geo_arr = np.zeros_like(y_m, dtype=float)

    for i, y in enumerate(y_mm):
        seg, t = _pick_segment(segs, float(y))
        re_loc = _lerp(seg.re0, seg.re1, t)
        a_geo_deg = _lerp(seg.a0, seg.a1, t)

        a0_0, al0_0 = _get_a0_alpha0_cached(seg.foil0, re_loc)
        a0_1, al0_1 = _get_a0_alpha0_cached(seg.foil1, re_loc)

        a0_arr[i] = _lerp(a0_0, a0_1, t)
        alpha0_arr[i] = _lerp(al0_0, al0_1, t)
        alpha_geo_arr[i] = np.deg2rad(a_geo_deg)

    y_out_m, alpha_i_rad, e = _llt_solve_halfspan(y_m, c_m, a0_arr, alpha_geo_arr, alpha0_arr)
    alpha_i_deg = np.rad2deg(alpha_i_rad)

    ws_base[e_cell].value = float(e)

    alpha_i_deg_on_y = np.interp(y_m, y_out_m, alpha_i_deg)

    for k, rr in enumerate(range(LIFT_A_START_ROW, LIFT_A_START_ROW + len(y_all_mm))):
        if k in idx_map:
            j = idx_map[k]
            ws_lift.cell(row=rr, column=INDUCED_COL).value = float(alpha_i_deg_on_y[j])
        else:
            ws_lift.cell(row=rr, column=INDUCED_COL).value = None

    for k, rr in enumerate(range(LIFT_A_START_ROW, LIFT_A_START_ROW + len(y_all_mm))):
        if k not in idx_map:
            ws_lift.cell(row=rr, column=OUT_CL_COL).value  = None
            ws_lift.cell(row=rr, column=OUT_CD_COL).value  = None
            ws_lift.cell(row=rr, column=OUT_CM_COL).value  = None
            ws_lift.cell(row=rr, column=OUT_XCP_COL).value = None
            continue

        j = idx_map[k]
        y = float(y_mm[j])
        seg, t = _pick_segment(segs, y)

        re_loc = _lerp(seg.re0, seg.re1, t)
        a_geo_deg = _lerp(seg.a0, seg.a1, t)
        alpha_eff = a_geo_deg - float(alpha_i_deg_on_y[j])

        polar0 = _get_polar_cached(seg.foil0, re_loc)
        polar1 = _get_polar_cached(seg.foil1, re_loc)

        cl0, cd0, cm0, xcp0 = _lookup_coeffs_from_polar(polar0, alpha_eff)
        cl1, cd1, cm1, xcp1 = _lookup_coeffs_from_polar(polar1, alpha_eff)

        ws_lift.cell(row=rr, column=OUT_CL_COL).value  = float(_lerp(cl0,  cl1,  t))
        ws_lift.cell(row=rr, column=OUT_CD_COL).value  = float(_lerp(cd0,  cd1,  t))
        ws_lift.cell(row=rr, column=OUT_CM_COL).value  = float(_lerp(cm0,  cm1,  t))
        ws_lift.cell(row=rr, column=OUT_XCP_COL).value = float(_lerp(xcp0, xcp1, t))

    print(
        f"[LLT] ({base_sheet_name} -> {lift_sheet_name}) done: e={e:.4f} -> {base_sheet_name}!{e_cell}, "
        f"induced alpha -> {lift_sheet_name} col {INDUCED_COL}, "
        f"coeffs -> {OUT_CL_COL}..{OUT_XCP_COL}, y_tip(LLT)={y_tip_mm:.1f}mm"
    )



def main() -> None:
    if not os.path.exists(WORKBOOK_PATH):
        raise FileNotFoundError(f"設計シートが見つかりません: {WORKBOOK_PATH}")
    if not os.path.isdir(CSV_FOLDER):
        raise NotADirectoryError(f"CSV フォルダが存在しません: {CSV_FOLDER}")

    wb = load_workbook(WORKBOOK_PATH)

    # ==========================================================
    # B案（揚力分布シートの chord を参照して Re を作る）では、
    # 先に chord を揚力分布シートへ書き込んでおくのが前提。
    # ==========================================================
    for base_name, lift_name in [
        (BASE_SHEET,  LIFT_SHEET),
        (BASE_SHEET2, LIFT_SHEET2),
    ]:
        if (base_name in wb.sheetnames) and (lift_name in wb.sheetnames):
            write_chord_to_lift_sheet(wb, base_name, lift_name)
        else:
            print(f"[WARN] chord書込みをスキップ: 必要なシートが見つかりません。 base={base_name}, lift={lift_name}")

    # ---- BASE/BASE限界 -> （削除済み）（Reは自動計算）----

    # ---- LLT & 揚力分布書き戻し（2系統）----
    for base_name, lift_name in [
        (BASE_SHEET,  LIFT_SHEET),
        (BASE_SHEET2, LIFT_SHEET2),
    ]:
        if (base_name in wb.sheetnames) and (lift_name in wb.sheetnames):
            llt_and_write_back(wb, base_name, lift_name, e_cell="N7")
        else:
            print(f"[WARN] LLTをスキップ: 必要なシートが見つかりません。 base={base_name}, lift={lift_name}")

    wb.save(WORKBOOK_PATH)
    print(f"[SAVE] 完了: {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
